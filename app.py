from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import load_workbook
from openpyxl.styles.numbers import is_date_format
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime
import os, uuid

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "output"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# --------------------------------------------------
# Detect header row (row with max filled cells)
# --------------------------------------------------
def detect_header_row(sheet):
    max_filled = 0
    header_row = 1

    for idx, row in enumerate(sheet.iter_rows(), start=1):
        filled = sum(1 for cell in row if cell.value not in (None, ""))
        if filled > max_filled:
            max_filled = filled
            header_row = idx

    return header_row


# --------------------------------------------------
# Exact Excel displayed value
# --------------------------------------------------
def excel_display_value(cell):
    """Return EXACT Excel displayed value"""
    if cell.value is None:
        return ""

    # Dates
    if is_date_format(cell.number_format):
        if isinstance(cell.value, datetime):
            return cell.value.strftime("%Y-%m-%d")

    # Numbers → convert using Excel formatting
    if isinstance(cell.value, (int, float)):
        text = format(cell.value, 'f')
        if '.' in text:
            text = text.rstrip('0').rstrip('.')
        return text

    # Everything else (text, etc.)
    return str(cell.value)


# --------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


# --------------------------------------------------
# Upload & preview
# --------------------------------------------------
@app.route("/upload", methods=["POST"])
def upload():
    file = request.files["file"]
    file_id = str(uuid.uuid4())

    path = os.path.join(UPLOAD_FOLDER, file_id + ".xlsx")
    file.save(path)

    wb = load_workbook(path, data_only=True)
    sheet = wb.active

    header_row = detect_header_row(sheet)

    headers = []
    rows = []

    for i, row in enumerate(sheet.iter_rows(), start=1):
        if i == header_row:
            headers = [excel_display_value(c) for c in row]
        elif i > header_row:
            rows.append([excel_display_value(c) for c in row])

    return jsonify({
        "headers": headers,
        "rows": rows
    })


# --------------------------------------------------
# Generate PDF
# --------------------------------------------------
@app.route("/generate_pdf", methods=["POST"])
def generate_pdf():
    data = request.json
    headers = data["headers"]
    rows = data["rows"]

    pdf_name = f"{uuid.uuid4()}.pdf"
    pdf_path = os.path.join(OUTPUT_FOLDER, pdf_name)

    styles = getSampleStyleSheet()
    styleN = styles["Normal"]

    table_data = [[Paragraph(h, styleN) for h in headers]]
    for row in rows:
        table_data.append([Paragraph(str(c), styleN) for c in row])

    # Auto column width
    col_widths = []
    for col in zip(*table_data):
        max_len = max(len(cell.text) for cell in col)
        col_widths.append(max(60, max_len * 6))

    page_width, _ = landscape(A4)
    usable_width = page_width - 40
    total_width = sum(col_widths)

    if total_width > usable_width:
        scale = usable_width / total_width
        col_widths = [w * scale for w in col_widths]

    pdf = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))

    pdf.build([table])

    return jsonify({"pdf": pdf_name})


# --------------------------------------------------
# Download PDF
# --------------------------------------------------
@app.route("/download/<filename>")
def download(filename):
    return send_file(
        os.path.join(OUTPUT_FOLDER, filename),
        as_attachment=True
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)

